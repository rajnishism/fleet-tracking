"""
export_to_excel.py
Export all Fleet Tracker database tables to an Excel workbook.

Exports data from both:
  - Backend database (backend/prisma/fleet.db): Vehicle, Location, Alert
  - Edge database (edge/database/vehicle_data.db): gps_records

Usage:
  python3 export_to_excel.py
"""

import sqlite3
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_DB = os.path.join(BASE_DIR, "backend", "prisma", "fleet.db")
EDGE_DB = os.path.join(BASE_DIR, "edge", "database", "vehicle_data.db")
OUTPUT_FILE = os.path.join(BASE_DIR, "fleet_tracker_export.xlsx")

# ─── Styling ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
HEADER_FILL_RED = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def export_table(wb, sheet_name, db_path, query, header_fill=HEADER_FILL):
    """Export a SQL query result to a styled Excel sheet."""
    if not os.path.exists(db_path):
        print(f"  ⚠️  Database not found: {db_path} — skipping {sheet_name}")
        return 0

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        cursor.execute(query)
    except Exception as e:
        print(f"  ⚠️  Query failed for {sheet_name}: {e}")
        conn.close()
        return 0

    rows = cursor.fetchall()
    if not rows:
        print(f"  ℹ️  No data in {sheet_name}")
        conn.close()
        return 0

    columns = [desc[0] for desc in cursor.description]
    conn.close()

    ws = wb.create_sheet(title=sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(dict(row).values(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Auto-width columns
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    return len(rows)


def main():
    print("📊 Fleet Tracker → Excel Export")
    print("=" * 40)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Backend tables
    count = export_table(
        wb, "Vehicles", BACKEND_DB,
        "SELECT vehicleId, name, createdAt FROM Vehicle ORDER BY vehicleId",
        header_fill=HEADER_FILL,
    )
    print(f"  ✅ Vehicles: {count} rows")

    count = export_table(
        wb, "Locations", BACKEND_DB,
        "SELECT id, vehicleId, latitude, longitude, elevation, speedKmh, timestamp, distFromRouteM, distanceStepM, fuelConsumedLiters, createdAt FROM Location ORDER BY id DESC",
        header_fill=HEADER_FILL_GREEN,
    )
    print(f"  ✅ Locations: {count} rows")

    count = export_table(
        wb, "Alerts", BACKEND_DB,
        "SELECT id, vehicleId, type, message, metadata, timestamp FROM Alert ORDER BY id DESC",
        header_fill=HEADER_FILL_RED,
    )
    print(f"  ✅ Alerts: {count} rows")

    # Edge table
    count = export_table(
        wb, "Edge GPS Records", EDGE_DB,
        "SELECT id, vehicle_id, latitude, longitude, speed_kmh, timestamp, distance_from_route_m, distance_step_m, fuel_consumed_liters, alerts, synced, created_at FROM gps_records ORDER BY id DESC",
        header_fill=HEADER_FILL_PURPLE,
    )
    print(f"  ✅ Edge GPS Records: {count} rows")

    # Save
    if wb.sheetnames:
        wb.save(OUTPUT_FILE)
        print(f"\n📁 Exported to: {OUTPUT_FILE}")
    else:
        print("\n⚠️  No data found in any table.")


if __name__ == "__main__":
    main()
